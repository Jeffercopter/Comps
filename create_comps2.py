import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EV-EBITDA Comps - Staffing"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
link_font = Font(color="0563C1", underline="single")
currency_fmt = '#,##0.0'
pct_fmt = '0%'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = [
    "Target", "Acquirer", "Percentage\nAcquired (%)", "Country",
    "Completion\nDate", "EBITDA\n(A$m)", "Enterprise\nValue (A$m)",
    "EV/EBITDA\nMultiple", "Source"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# All 29 transactions
# Format: (target, acquirer, pct, country, date, ebitda, ev, multiple_formula_or_str, source_url)
# ebitda/ev: number or None for "Not Disclosed"
# multiple: "formula" for =Gx/Fx, or None for "Not Disclosed"
transactions = [
    ("Infrawork Holdings Limited",
     "Peoplein Limited", 1.0, "New Zealand", "Feb-26",
     4.6, 22.0, True,
     "https://www.nzherald.co.nz/business/economy/employment/infrawork-sold-to-asx-listed-peoplein-in-56m-immigration-labour-deal/F54YAHVVZBC57K4E5MHVCI54OQ/"),

    ("First Choice Care/Edmen",
     "Healthcare Australia Pty Ltd.", 1.0, "Australia", "Dec-25",
     3.27, 20.25, True,
     "https://www.businessnewsaustralia.com/articles/peoplein-to-sell-first-choice-care-and-edmen.html"),

    ("Techforce Personnel Pty Limited",
     "ina.", 0.79, "Australia", "Dec-25",
     None, 23.5, None,
     "https://www.listcorp.com/asx/ppe/peoplein-limited/news/peoplein-to-divest-techforce-personnel-3283656.html"),

    ("APM Human Services International Limited",
     "Madison Dearborn Partners, LLC", 0.71, "Australia", "Sep-24",
     281.0, 2100.0, True,
     "https://www.marketscreener.com/quote/stock/APM-HUMA-129330189/news/Madison-Dearborn-Partners-LLC-completed-the-acquisition-of-remaining-71-stake-in-APM-Human-Service-47940865/"),

    ("Owen Pacific Workforce Pty Ltd",
     "Ashley Services Group Limited", 1.0, "Australia", "Feb-23",
     4.5, 15.0, True,
     "https://newsnreleases.com/2023/02/03/ashley-services-group-acquires-owen-pacific-workforce-for-up-to-15-9-million/"),

    ("Arbor E&T, LLC (dba Equus Workforce Solutions)",
     "APM Human Services International Limited", 1.0, "United States", "Nov-22",
     50.0, 239.8, True,
     "https://www.listcorp.com/asx/apm/apm-human-services-international/news/apm-completes-acquisition-of-equus-workforce-solutions-2792468.html"),

    ("Linc Personnel Pty Ltd",
     "Ashley Services Group Limited", 0.75, "Australia", "Jul-22",
     1.4, 5.6, True,
     "https://www.listcorp.com/asx/ash/ashley-services-group/news/ashley-acquires-major-shareholding-in-linc-personnel-2730158.html"),

    ("FIP Group Holdings Pty Ltd",
     "Peoplein Limited", 1.0, "Australia", "Jun-22",
     9.5, 45.0, True,
     "https://www.moneymorning.com.au/20220603/peoplein-asxppe-poised-to-take-fip-as-shares-spike.html"),

    ("Perigon Group Pty Ltd",
     "Peoplein Limited", 1.0, "Australia", "Feb-22",
     4.3, 16.0, True,
     "https://themarketherald.com.au/peoplein-asxppe-to-acquire-perigon-group-2022-02-09/"),

    ("Clustera Sverige AB",
     "APM Human Services International Limited", 1.0, "Sweden", "Jan-22",
     2.0, 15.0, True,
     "https://www.listcorp.com/asx/apm/apm-human-services-international/news/apm-completes-ipo-acquisitions-and-amp-swedish-market-entry-2652618.html"),

    ("GMT Group Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Dec-21",
     None, None, None,
     "https://www.staffingindustry.com/news/global-daily-news/australia-people-infrastructure-completes-acquisition-gmt-people"),

    ("Inverse Group Pty. Ltd.",
     "Hiremii Limited", 1.0, "Australia", "Nov-21",
     None, 1.5, None,
     "https://www.proactiveinvestors.com/companies/news/965371/hiremii-s-inverse-acquisition-to-increase-revenue-client-base-and-footprint-965371.html"),

    ("Vision Surveys (QLD) Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 0.75, "Australia", "Jul-21",
     None, 6.7, None,
     "https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-completed-the-acquisition-of-75-majority-stake-in-Vision-Surveys-Pty-Ltd-36025964/"),

    ("Techforce Personnel Limited",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 0.79, "Australia", "Jun-21",
     None, 30.0, None,
     "https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-entered-into-binding-agreements-to-acquire-Techforce-Personnel-Pty-Ltd-for-35497394/"),

    ("Business of SwingShift Nurses",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Mar-21",
     1.0, 3.1, True,
     "https://themarketherald.com.au/people-infrastructure-asxppe-purchases-swingshift-nurses-for-3-1m-2021-03-15/"),

    ("Oncontractor Pty. Ltd.",
     "Hiremii Limited", 1.0, "Australia", "Feb-21",
     None, None, None,
     "https://www.proactiveinvestors.com/companies/news/960608/hiremii-well-funded-to-advance-ai-cloud-recruitment-platform-960608.html"),

    ("Assets of ECT4Health Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Feb-21",
     None, None, None,
     "https://announcements.asx.com.au/asxpdf/20210826/pdf/44zrh6npzvxvmm.pdf"),

    ("Ecareer Employment Services Pty Ltd / Illuminate Search and Consulting",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Jan-21",
     1.3, 5.2, True,
     "https://www.listcorp.com/asx/ppe/people-infrastructure-ltd/news/completion-of-ecareers-and-illuminate-acquistions-2499487.html"),

    ("The Instruction Company Pty Ltd",
     "Ashley Services Group Limited", 1.0, "Australia", "Sep-20",
     None, None, None,
     "https://ashleyservicesgroup.com.au/investor-centre/"),

    ("Serendipity (WA) Pty Ltd (t/a APM)",
     "Madison Dearborn Partners, LLC", 0.55, "Australia", "Jun-20",
     148.4, 1600.0, True,
     "https://michaelwest.com.au/apm-promoters-exit-jobless-profits-in-asx-float-turn-to-profiteering-from-disabled/"),

    ("Ccl Group Holdings Pty. Ltd.",
     "Ashley Services Group Limited", 0.80, "Australia", "Dec-19",
     4.1, 14.0, True,
     "https://www.staffingindustry.com/news/global-daily-news/australia-ashley-services-acquires-ccl-group"),

    ("First Choice Care Pty. Ltd.",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Jul-19",
     None, 11.0, None,
     "https://www.businessnewsaustralia.com/articles/people-infrastructure-acquires-nursing-recruitment-firms.html"),

    ("Carestaff Healthcare Solutions Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Jun-19",
     None, 5.8, None,
     "https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-completed-the-acquisition-of-Carestaff-Nursing-Services-Pty-Ltd-for-AUD-5-34255922/"),

    ("Halcyon Knights / Halcyon Knights Commercial and Contracting",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Jun-19",
     3.9, 13.5, True,
     "https://www.businessnewsaustralia.com/articles/people-infrastructure-buys-halcyon-knights-for--13-5m.html"),

    ("Remaining Shares in Recon Solutions Pty Ltd and Project Partners Corp",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Mar-19",
     None, 2.8, None,
     "https://www.businessnewsaustralia.com/articles/steve-scanlan-sells-remaining-stake-in-recon-group.html"),

    ("Victorian Nurse Specialists Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Mar-19",
     None, 2.5, None,
     "https://www.finnewsnetwork.com.au/archives/finance_news_network222179.html"),

    ("Network Nursing Agency Pty Ltd / Australian Healthcare Academy Pty Ltd",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Aug-18",
     2.65, 8.0, True,
     "https://www.staffingindustry.com/news/global-daily-news/australia-people-infrastructure-acquires-nursing-agency-business"),

    ("50% Stake in Recon Solutions Pty Ltd and 50% Stake in Recon Technology Solutions",
     "People Infrastructure Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Jan-18",
     0.8, 2.8, True,
     "https://www.proactiveinvestors.com.au/companies/news/189851/people-infrastructure-s-acquisition-will-drive-revenue-higher-189851.html"),

    ("AWX Pty Ltd",
     "People Infrastructure Pty Ltd (nka: Peoplein Limited)", 1.0, "Australia", "Oct-16",
     None, 18.4, None,
     "https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-acquired-AWX-Pty-Ltd-for-AUD-18-4-million-35119752/"),
]

for i, txn in enumerate(transactions):
    row = i + 2
    target, acquirer, pct, country, date, ebitda, ev, mult, source = txn

    ws.cell(row=row, column=1, value=target).alignment = left_align
    ws.cell(row=row, column=2, value=acquirer).alignment = left_align

    pct_cell = ws.cell(row=row, column=3, value=pct)
    pct_cell.number_format = pct_fmt
    pct_cell.alignment = data_align

    ws.cell(row=row, column=4, value=country).alignment = data_align
    ws.cell(row=row, column=5, value=date).alignment = data_align

    # EBITDA
    ebitda_cell = ws.cell(row=row, column=6)
    if ebitda is None:
        ebitda_cell.value = "Not Disclosed"
    else:
        ebitda_cell.value = ebitda
        ebitda_cell.number_format = currency_fmt
    ebitda_cell.alignment = data_align

    # EV
    ev_cell = ws.cell(row=row, column=7)
    if ev is None:
        ev_cell.value = "Not Disclosed"
    else:
        ev_cell.value = ev
        ev_cell.number_format = currency_fmt
    ev_cell.alignment = data_align

    # Multiple
    mult_cell = ws.cell(row=row, column=8)
    if mult is True:
        mult_cell.value = f"=G{row}/F{row}"
        mult_cell.number_format = '0.0"x"'
    else:
        mult_cell.value = "Not Disclosed"
    mult_cell.alignment = data_align

    # Source
    source_cell = ws.cell(row=row, column=9)
    source_cell.value = source
    source_cell.hyperlink = source
    source_cell.font = link_font
    source_cell.alignment = left_align

    for col in range(1, 10):
        ws.cell(row=row, column=col).border = thin_border

# Summary stats
last_data_row = len(transactions) + 1
sr = last_data_row + 2
ws.cell(row=sr, column=1, value="Summary Statistics").font = Font(bold=True)

ws.cell(row=sr+1, column=7, value="Mean EV/EBITDA:").font = Font(bold=True)
ws.cell(row=sr+1, column=7).alignment = Alignment(horizontal="right")
mean_cell = ws.cell(row=sr+1, column=8)
mean_cell.value = f"=AVERAGE(H2:H{last_data_row})"
mean_cell.number_format = '0.0"x"'
mean_cell.font = Font(bold=True)
mean_cell.alignment = data_align

ws.cell(row=sr+2, column=7, value="Median EV/EBITDA:").font = Font(bold=True)
ws.cell(row=sr+2, column=7).alignment = Alignment(horizontal="right")
median_cell = ws.cell(row=sr+2, column=8)
median_cell.value = f"=MEDIAN(H2:H{last_data_row})"
median_cell.number_format = '0.0"x"'
median_cell.font = Font(bold=True)
median_cell.alignment = data_align

# Notes
nr = sr + 4
notes = [
    "Notes:",
    "1. All figures in AUD millions unless otherwise stated.",
    "2. Infrawork Holdings: NZD-denominated deal (NZD 56m total / NZD 24m upfront). AUD figures approximate at ~A$22m upfront EV / A$4.6m EBITDA. Multiple of 4.8x is on upfront basis. Source: PeopleIN H1 FY26 earnings call.",
    "3. First Choice Care/Edmen (Dec-25): PeopleIN stated sale at '6.2 times FY25 earnings'. Exact earnings metric (EBITDA vs EBITA vs NPAT) not specified.",
    "4. Techforce Personnel (Dec-25): Divestment by PeopleIN for A$23.5m (79.3% stake). No EBITDA or multiple publicly disclosed.",
    "5. APM Human Services: EV implied from equity value (~A$1.33bn at $1.45/share) + net debt (~A$800m). FY24 underlying EBITDA of A$281m.",
    "6. Owen Pacific Workforce: Consideration range A$14.2-15.9m based on FY23 EBITDA outcome (A$4.1-5.0m). Mid-point used.",
    "7. Linc Personnel: A$4.2m for 75% stake; implied 100% EV ~A$5.6m. Normalised FY22 EBITDA of A$1.4m.",
    "8. FIP Group: A$45m upfront (A$70m max incl. earnout). FY23 pro forma EBITDA A$9.5m. Multiple shown is on upfront basis.",
    "9. Perigon Group: A$16m upfront (A$26.8m max incl. earnout). FY23E EBITDA A$4.3m. Multiple shown is on upfront basis.",
    "10. Clustera Sverige: SEK 100m (~A$15m). FY22 adjusted forecast EBITDA SEK 12m (~A$2m).",
    "11. GMT Group: Financial terms not publicly disclosed; described as immaterial to FY21 earnings.",
    "12. Inverse Group: Max consideration A$1.5m; structured on revenue milestones, not EBITDA.",
    "13. Vision Surveys & Techforce Personnel (Jun/Jul-21): Combined FY22E EBITDA of A$5.8m disclosed; individual EBITDA not separately available.",
    "14. Oncontractor: All-scrip deal (corporate restructuring); no EV/EBITDA disclosed.",
    "15. ECT4Health: Small asset acquisition; financial terms not considered material for disclosure.",
    "16. Ecareer/Illuminate: A$5.2m net of cash. Forward 12-month EBITDA A$1.3m.",
    "17. The Instruction Company: Financial terms not publicly disclosed.",
    "18. Serendipity (WA) / APM: 55% stake acquired by MDP from Quadrant PE. EV ~A$1.6bn, FY20 EBITDA A$148.4m. Large diversified human services platform, not pure staffing.",
    "19. CCL Group: 80% stake; implied 100% EV ~A$14m. Normalised FY19 EBITDA A$4.1m.",
    "20. First Choice Care (Jul-19) & Carestaff (Jun-19): Acquired together for combined A$16.8m / combined expected EBITDA A$3.4m (~4.9x combined). Individual splits not disclosed.",
    "21. Halcyon Knights: A$13.5m upfront (A$21.75m max incl. earnout). FY20E EBITDA A$3.9m. Multiple shown is on upfront basis.",
    "22. Network Nursing Agency: A$8.0m upfront (A$9.1m max). FY19 earn-out threshold EBITDA ~A$2.65m used as proxy.",
    "23. 50% Recon Solutions/Recon Technology (Jan-18): A$2.8m for 50% stake. PPE's 50% share EBITDA contribution A$0.8m.",
    "24. AWX: Pre-IPO acquisition for A$18.4m (A$16.6m cash + A$1.8m earnout). EBITDA not publicly available.",
]

for j, note in enumerate(notes):
    cell = ws.cell(row=nr+j, column=1, value=note)
    cell.font = Font(italic=True) if j > 0 else Font(bold=True, italic=True)

# Column widths
from openpyxl.utils import get_column_letter
col_widths = {1: 55, 2: 55, 3: 14, 4: 16, 5: 14, 6: 14, 7: 18, 8: 16, 9: 90}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 35

wb.save("/home/user/Comps/EV_EBITDA_Comps_Staffing.xlsx")
print("Excel file created successfully!")
