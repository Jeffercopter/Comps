"""Syama_Pebble_Sorting_Simulation.xlsx
STEINERT Sukari model SR241558 re-parameterised for the Syama converted SAG line."""
import openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter as CL

OUT = sys.argv[1]
NAVY, STEEL, LIGHT, GREY, AMBER, GREEN = "1F3864","2E5C8A","D9E2F3","F2F2F2","FFF2CC","E2EFDA"
H1 = Font(size=14, bold=True, color="FFFFFF"); HS = Font(size=10, bold=True, color="FFFFFF")
BD = Font(size=10, bold=True); NM = Font(size=10)
IT = Font(size=9, italic=True, color="595959"); KEY = Font(size=10, bold=True, color="C00000")
fill = lambda c: PatternFill("solid", fgColor=c)
_t = Side(style="thin", color="BFBFBF"); BOX = Border(_t,_t,_t,_t)

def put(ws, cell, v, font=NM, num=None, f=None, al=None, bd=True):
    c = ws[cell]; c.value = v; c.font = font
    if num: c.number_format = num
    if f: c.fill = fill(f)
    if al: c.alignment = Alignment(horizontal=al)
    if bd: c.border = BOX
    return c

def band(ws, r, c1, c2, text, f=STEEL, font=HS):
    ws.cell(r, c1).value = text
    for c in range(c1, c2+1):
        ws.cell(r, c).font = font; ws.cell(r, c).fill = fill(f)

wb = openpyxl.Workbook()

# --------------------------------------------------- fixed mass-balance geometry
BLK = [10 + 14*i for i in range(6)]     # pass blocks
T   = 98                                # circuit-total band
R = dict(mill=T+1, screen=T+4, over=T+5, under=T+6, sfeed=T+8, rej=T+9, con=T+10)
PERF = T + 13

# ====================================================================== DASHBOARD
d = wb.active; d.title = "Dashboard"; d.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLMN", [2,36,12,8,2,46,19,2,31,12,12,11,8,32]):
    d.column_dimensions[col].width = w
band(d, 1, 1, 14, "  SYAMA GOLD  |  SAG PEBBLE (SCAT) SORTING  |  BUSINESS CASE", NAVY, H1)
put(d,"A2","Project:",BD,bd=False); put(d,"B2","Syama Sulphide Conversion Project - converted (ex-oxide) SAG line",bd=False)
put(d,"A3","Basis:",BD,bd=False);   put(d,"B3","STEINERT model SR241558 (Sukari) re-parameterised for Syama",bd=False)
put(d,"A4","Date:",BD,bd=False);    put(d,"B4","2026-08-05",bd=False)
band(d, 6, 2, 7, "  INPUTS"); band(d, 6, 9, 14, "  OUTPUTS")
for c,t in [("B7","Parameter"),("C7","Value"),("D7","Unit"),("F7","Basis / source"),("G7","Confidence")]:
    put(d,c,t,BD,f=LIGHT)
for c,t in [("I7","Parameter"),("J7","BASE"),("K7","SORTING"),("L7","DIFF"),("M7","Unit"),("N7","Comment")]:
    put(d,c,t,BD,f=LIGHT,al="center" if c in ("J7","K7","L7") else None)

IN = [
 ("§","OPERATION",None,None,None,None),
 ("feed","New feed to SAG line",200.0,"t/h","1.58 Mtpa - the SSCP increment (2.4 -> 4.0 Mtpa)","Assumption - CONFIRM","#,##0"),
 ("hrs","Annual operating hours",7884.0,"h/a","90% availability","Assumption","#,##0"),
 ("rom","ROM grade (diluted)",2.40,"g/t","2026 UG plan 2.4-2.5 g/t Au","Resolute guidance","0.00"),
 ("rec","Gold recovery (flot+roast+CIL)",0.75,"-","FY25 sulphide 75%, Q2-25 76%","Resolute actual","0.0%"),
 ("poz","Gold price",4000.0,"$/oz","Resolute 2026 planning assumption","Resolute","#,##0"),
 ("pg","Gold price",None,"$/g","gold price / 31.1035","Calculated","0.00"),
 ("disc","Discount rate",0.10,"-","8% base + Mali country-risk premium","Assumption","0.0%"),
 ("§","COSTS",None,None,None,None),
 ("cmill","Milling cost avoided",8.00,"$/t","power + media + reagents (Sukari used 7.69)","Assumption - CONFIRM","0.00"),
 ("cmine","Incremental UG mining cost",53.00,"$/t","~45% of AISC $2,050/oz at 1.8 g/t recovered","Derived - CONFIRM","0.00"),
 ("cdump","Reject rehandle + dump",2.50,"$/t","","Assumption","0.00"),
 ("csort","Sorting OPEX",0.45,"$/t","STEINERT template ($15/50 t/h + $0.15)","Vendor benchmark","0.00"),
 ("§","MATERIAL PROPERTY",None,None,None,None),
 ("dil","Waste dilution (SLC)",0.20,"-","sublevel cave, typical 15-25%","Assumption - CONFIRM","0.0%"),
 ("og","Undiluted ore grade",None,"g/t","ROM grade / (1 - dilution)","Calculated","0.000"),
 ("fo","Mass yield ore -> scats",0.1111,"-","calibrated to the measured scat stream","Calibrated","0.0000"),
 ("fw","Mass yield waste -> scats",0.2000,"-","calibrated to the measured scat stream","Calibrated","0.0000"),
 ("con","HETEROGENEITY CONTRAST",None,"x","waste yield / ore yield.   Sukari = 8.9x","KEY METRIC","0.00"),
 ("§","SORTER PROPERTY",None,None,None,None),
 ("orec","Sorter ore recovery",0.95,"-","XRT, conservatively tuned","TEST WORK REQUIRED","0.0%"),
 ("w2r","Waste mass to rejects",0.70,"-","XRT on Syama lithologies","TEST WORK REQUIRED","0.0%"),
 ("§","CAPITAL & EVALUATION",None,None,None,None),
 ("capex","Sorting plant CAPEX",4095000.0,"$","1 x 50 t/h sorter, STEINERT build-up +30%","Vendor benchmark","#,##0"),
 ("build","Build period",12,"months","","Assumption","0"),
 ("life","Evaluation life",96,"months","8 years","Assumption","0"),
 ("upl","ROM feed uplift  (Goal Seek)",7.2,"t/h",'Goal Seek: set L{under} to 0 by changing this cell',"Solved","0.00"),
 ("cap","Capacity capture factor",1.00,"-","share of freed SAG capacity actually filled with ore","CRITICAL ASSUMPTION","0.0%"),
]
r = 9; X = {}
for row in IN:
    if row[0] == "§":
        band(d, r, 1, 7, "  " + row[1]); r += 1; continue
    k, name, val, unit, basis, conf, nf = row
    put(d, f"B{r}", name, KEY if conf == "KEY METRIC" else BD)
    put(d, f"C{r}", val, NM, num=nf,
        f=AMBER if conf in ("Calibrated","CRITICAL ASSUMPTION","KEY METRIC","TEST WORK REQUIRED") else None)
    put(d, f"D{r}", unit); put(d, f"F{r}", basis, IT); put(d, f"G{r}", conf, IT)
    X[k] = r; r += 1
d[f"C{X['pg']}"]  = f"=C{X['poz']}/31.1035"
d[f"C{X['og']}"]  = f"=C{X['rom']}/(1-C{X['dil']})"
d[f"C{X['con']}"] = f"=C{X['fw']}/C{X['fo']}"

# calibration check
CB = r + 1
band(d, CB-1, 2, 7, "  CALIBRATION CHECK  -  model must reproduce the measured scat stream")
for i,(lbl,v,nf,conf) in enumerate([
        ("Measured circulating scat load", 30.0, "0.0", "CLIENT DATA"),
        ("MEASURED SCAT GRADE", 2.00, "0.00", "CLIENT DATA"),
        ("Modelled circulating scat load", f"='MB Base'!G{R['sfeed']}", "0.0", "calc"),
        ("Modelled scat grade", f"='MB Base'!H{R['sfeed']}", "0.00", "calc")]):
    put(d,f"B{CB+i}",lbl,BD); put(d,f"C{CB+i}",v,num=nf,f=AMBER if conf=="CLIENT DATA" else None)
    put(d,f"D{CB+i}","t/h" if i in (0,2) else "g/t")
    put(d,f"F{CB+i}","plant survey / client sampling" if conf=="CLIENT DATA"
        else "adjust the two mass yields above until modelled = measured", IT)
    put(d,f"G{CB+i}",conf,IT)
d[f"C{X['upl']}"] = 7.2
d[f"F{X['upl']}"] = f"Goal Seek: set L{9+8} to 0 by changing this cell"

# outputs
OUTP = [("ROM feed rate", f"=C{X['feed']}", "='MB Sort'!G5", "t/h", "uplift from rejecting scat mass"),
 ("ROM grade","='MB Base'!H5","='MB Sort'!H5","g/t",""),
 ("SAG mill total load",f"='MB Base'!G{R['mill']}",f"='MB Sort'!G{R['mill']}","t/h","new feed + circulating scats"),
 ("Scat / sorter feed",f"='MB Base'!G{R['sfeed']}",f"='MB Sort'!G{R['sfeed']}","t/h",""),
 ("Scat grade",f"='MB Base'!H{R['sfeed']}",f"='MB Sort'!H{R['sfeed']}","g/t","measured = 2.00 g/t"),
 ("Sorter rejects",f"='MB Base'!G{R['rej']}",f"='MB Sort'!G{R['rej']}","t/h","to waste dump"),
 ("Reject grade",f"='MB Base'!H{R['rej']}",f"='MB Sort'!H{R['rej']}","g/t","gold discarded"),
 ("Mass pull to rejects",f"=IFERROR('MB Base'!G{R['rej']}/'MB Base'!G{R['sfeed']},0)",
                         f"=IFERROR('MB Sort'!G{R['rej']}/'MB Sort'!G{R['sfeed']},0)","-","of the scat stream"),
 ("Ball-mill / leach feed",f"='MB Base'!G{R['under']}",f"='MB Sort'!G{R['under']}","t/h","HELD CONSTANT (Goal Seek)"),
 ("Leach feed grade",f"='MB Base'!H{R['under']}",f"='MB Sort'!H{R['under']}","g/t","the uplift"),
 ("Au to leach",f"='MB Base'!I{R['under']}",f"='MB Sort'!I{R['under']}","g/h",""),
 ("Au production",None,None,"oz/a","")]
r = 9; O = {}
for name, jf, kf, unit, note in OUTP:
    put(d,f"I{r}",name,BD); put(d,f"J{r}",jf,num="#,##0.000"); put(d,f"K{r}",kf,num="#,##0.000")
    put(d,f"L{r}",f"=K{r}-J{r}",num="#,##0.000"); put(d,f"M{r}",unit); put(d,f"N{r}",note,IT)
    O[name] = r; r += 1
for c in "JKL":
    d[f"{c}{O['Mass pull to rejects']}"].number_format = "0.0%"
    d[f"{c}{O['Au production']}"] = (f"={c}{O['Au to leach']}*$C${X['rec']}*$C${X['hrs']}/31.1035"
                                     if c != "L" else f"=K{O['Au production']}-J{O['Au production']}")
    d[f"{c}{O['Au production']}"].number_format = "#,##0"

r += 1
band(d, r, 9, 14, "  ECONOMICS   (sorting case less base case)"); r += 1
EC = [("Incremental gold to leach", f"=L{O['Au to leach']}", "g/h", "#,##0.0"),
      ("Incremental gold recovered", f"=L{O['Au production']}", "oz/a", "#,##0"),
      ("Revenue gain", f"=L{O['Au to leach']}*C{X['rec']}*C{X['pg']}", "$/h", "#,##0"),
      ("Sorting OPEX", f"=-C{X['csort']}*'MB Sort'!G{R['sfeed']}", "$/h", "#,##0"),
      ("Reject rehandle + dump", f"=-C{X['cdump']}*'MB Sort'!G{R['rej']}", "$/h", "#,##0"),
      ("Incremental mining cost", f"=-C{X['cmine']}*L{O['ROM feed rate']}", "$/h", "#,##0"),
      ("Milling cost saved", f"=-C{X['cmill']}*L{O['SAG mill total load']}", "$/h", "#,##0")]
E0 = r
for name, f_, unit, nf in EC:
    put(d,f"I{r}",name,BD); put(d,f"K{r}",f_,num=nf); put(d,f"M{r}",unit); r += 1
put(d,f"I{r}","NET CASH BENEFIT",KEY); put(d,f"K{r}",f"=SUM(K{E0+2}:K{r-1})",BD,num="#,##0",f=GREEN)
put(d,f"M{r}","$/h"); NETH = r; r += 1
put(d,f"I{r}","NET CASH BENEFIT",KEY); put(d,f"K{r}",f"=K{NETH}*C{X['hrs']}",BD,num="#,##0",f=GREEN)
put(d,f"M{r}","$/a"); NETA = r; r += 1
put(d,f"I{r}","NPV",KEY); put(d,f"K{r}","=NPV!B8",BD,num="#,##0",f=GREEN); put(d,f"M{r}","$"); r += 1
put(d,f"I{r}","IRR",KEY); put(d,f"K{r}","=NPV!B9",BD,num="0.0%",f=GREEN); put(d,f"M{r}","-"); r += 1
put(d,f"I{r}","Simple payback",KEY)
put(d,f"K{r}",f'=IF(K{NETA}>0,C{X["capex"]}/K{NETA}*12,"n/a")',BD,num="0.0",f=GREEN); put(d,f"M{r}","months")

# =============================================================== MASS BALANCE
COLS = dict(G="Mass t/h", H="Au g/t", I="Au g/h", J="Ore %", K="Ore t/h",
            L="Ore g/t", M="Ore Au g/h", N="Waste %", O="Waste t/h")

def mb(name, orec, w2r, feed_ref, subtitle):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLMNO", [2,32,11,7,2,30,11,10,11,9,10,10,11,9,10]):
        ws.column_dimensions[col].width = w
    band(ws, 1, 2, 15, f"  {name}   |   {subtitle}")
    for c,t in [("B2","Assumption"),("C2","Value"),("D2","Unit"),("F2","Stream")]: put(ws,c,t,BD,f=LIGHT)
    for c,t in COLS.items(): put(ws, f"{c}2", t, BD, f=LIGHT, al="center")
    for i,(lbl,f_,nf,u) in enumerate([
            ("Mass yield ore -> oversize",  f"=Dashboard!C{X['fo']}", "0.0000",""),
            ("Mass yield waste -> oversize",f"=Dashboard!C{X['fw']}", "0.0000",""),
            ("Sorter ore recovery",         orec, "0.0%",""),
            ("Waste mass to rejects",       w2r,  "0.0%",""),
            ("Undiluted ore grade",         f"=Dashboard!C{X['og']}", "0.000","g/t"),
            ("Waste dilution",              f"=Dashboard!C{X['dil']}","0.0%","")]):
        put(ws,f"B{3+i}",lbl); put(ws,f"C{3+i}",f_,num=nf); put(ws,f"D{3+i}",u)
    def stream(r, label, K, O, bold=False):
        put(ws,f"F{r}",label, BD if bold else NM)
        put(ws,f"K{r}",K,num="#,##0.000"); put(ws,f"O{r}",O,num="#,##0.000")
        put(ws,f"L{r}","=$C$7",num="#,##0.000"); put(ws,f"M{r}",f"=K{r}*L{r}",num="#,##0.000")
        put(ws,f"G{r}",f"=K{r}+O{r}",num="#,##0.000"); put(ws,f"I{r}",f"=M{r}",num="#,##0.000")
        put(ws,f"H{r}",f"=IFERROR(I{r}/G{r},0)",num="#,##0.000")
        put(ws,f"J{r}",f"=IFERROR(K{r}/G{r},0)",num="0.0%"); put(ws,f"N{r}",f"=IFERROR(O{r}/G{r},0)",num="0.0%")
    stream(5, "PLANT FEED (new ROM)", f"={feed_ref}*(1-$C$8)", f"={feed_ref}*$C$8", True)
    for c in COLS: ws[f"{c}5"].fill = fill(LIGHT)
    for i, r0 in enumerate(BLK):
        band(ws, r0, 6, 15, f"  RECIRCULATION PASS {i+1}", LIGHT, BD)
        src = "5" if i == 0 else str(BLK[i-1] + 10)
        stream(r0+1, "SAG feed", f"=K{src}", f"=O{src}")
        stream(r0+2, "SAG discharge", f"=K{r0+1}", f"=O{r0+1}")
        put(ws,f"F{r0+3}","Mill discharge screen",BD,f=GREY)
        stream(r0+4, "   screen feed", f"=K{r0+2}", f"=O{r0+2}")
        stream(r0+5, "   oversize (scats)", f"=K{r0+4}*$C$3", f"=O{r0+4}*$C$4")
        stream(r0+6, "   undersize", f"=K{r0+4}-K{r0+5}", f"=O{r0+4}-O{r0+5}")
        put(ws,f"F{r0+7}","Pebble sorter",BD,f=GREY)
        stream(r0+8, "   sorter feed", f"=K{r0+5}", f"=O{r0+5}")
        stream(r0+9, "   rejects", f"=K{r0+8}*(1-$C$5)", f"=O{r0+8}*$C$6")
        stream(r0+10,"   concentrate -> SAG", f"=K{r0+8}-K{r0+9}", f"=O{r0+8}-O{r0+9}")
    band(ws, T, 6, 15, "  CIRCUIT TOTAL")
    sK = "+".join(f"K{s+1}" for s in BLK); sO = "+".join(f"O{s+1}" for s in BLK)
    stream(R['mill'],   "SAG MILL FEED (total load)", f"={sK}", f"={sO}", True)
    stream(R['mill']+1, "SAG mill discharge", f"=K{R['mill']}", f"=O{R['mill']}")
    put(ws,f"F{R['mill']+2}","Mill discharge screen",BD,f=GREY)
    stream(R['screen'], "   screen feed", f"=K{R['mill']+1}", f"=O{R['mill']+1}")
    stream(R['over'],   "   oversize (scats)", f"=K{R['screen']}*$C$3", f"=O{R['screen']}*$C$4")
    stream(R['under'],  "   UNDERSIZE -> ball mill / leach", f"=K{R['screen']}-K{R['over']}",
           f"=O{R['screen']}-O{R['over']}", True)
    put(ws,f"F{R['sfeed']-1}","Pebble sorter",BD,f=GREY)
    stream(R['sfeed'], "   SORTER FEED", f"=K{R['over']}", f"=O{R['over']}", True)
    stream(R['rej'],   "   REJECTS -> waste dump", f"=K{R['sfeed']}*(1-$C$5)", f"=O{R['sfeed']}*$C$6", True)
    stream(R['con'],   "   CONCENTRATE -> SAG", f"=K{R['sfeed']}-K{R['rej']}", f"=O{R['sfeed']}-O{R['rej']}", True)
    for rr, col in ((R['under'], GREEN), (R['sfeed'], AMBER), (R['rej'], AMBER)):
        for c in COLS: ws[f"{c}{rr}"].fill = fill(col)
    band(ws, PERF, 2, 8, "  SORTER PERFORMANCE")
    for i,(lbl,f_,nf,u) in enumerate([
        ("Mass pull to rejects", f"=IFERROR(G{R['rej']}/G{R['sfeed']},0)","0.0%",""),
        ("Au recovery to concentrate", f"=IFERROR(I{R['con']}/I{R['sfeed']},1)","0.0%",""),
        ("Upgrade ratio (conc / feed)", f"=IFERROR(H{R['con']}/H{R['sfeed']},1)","0.000","x"),
        ("Au lost to rejects", f"=I{R['rej']}","#,##0.00","g/h"),
        ("Recoverable value discarded", f"=I{R['rej']}*Dashboard!C{X['rec']}*Dashboard!C{X['pg']}","#,##0","$/h"),
        ("In-situ value of reject stream", f"=H{R['rej']}*Dashboard!C{X['rec']}*Dashboard!C{X['pg']}","#,##0.00","$/t"),
        ("Cost avoided per reject tonne", f"=Dashboard!C{X['cmill']}-Dashboard!C{X['cdump']}","#,##0.00","$/t")]):
        put(ws,f"B{PERF+1+i}",lbl,BD); put(ws,f"C{PERF+1+i}",f_,num=nf); put(ws,f"D{PERF+1+i}",u)
    return ws

mb("MB Base", 1.0, 0.0, f"Dashboard!C{X['feed']}",
   "base case - all scats crushed and returned, no sorting")
mb("MB Sort", f"=Dashboard!C{X['orec']}", f"=Dashboard!C{X['w2r']}",
   f"(Dashboard!C{X['feed']}+Dashboard!C{X['upl']}*Dashboard!C{X['cap']})",
   "sorting case - barren scat mass rejected, freed SAG capacity refilled with ROM")

# ============================================================================ NPV
n = wb.create_sheet("NPV"); n.sheet_view.showGridLines = False
n.column_dimensions["A"].width = 32; n.column_dimensions["B"].width = 16; n.column_dimensions["C"].width = 14
band(n, 1, 1, 8, "  NPV   |   monthly cash flow")
for i,(lbl,f_,nf,u) in enumerate([
        ("CAPEX", f"=Dashboard!C{X['capex']}", "#,##0","$"),
        ("Net benefit", f"=Dashboard!K{NETA}", "#,##0","$/a"),
        ("Discount rate", f"=Dashboard!C{X['disc']}", "0.0%",""),
        ("Build period", f"=Dashboard!C{X['build']}", "0","months"),
        ("Evaluation life", f"=Dashboard!C{X['life']}", "0","months")]):
    put(n,f"A{3+i}",lbl,BD); put(n,f"B{3+i}",f_,num=nf); put(n,f"C{3+i}",u,bd=False)
NM_ = 120
last = CL(3 + NM_)
put(n,"A8","NPV",KEY); put(n,"B8",f"=NPV(B5/12,D13:{last}13)",BD,num="#,##0",f=GREEN); put(n,"C8","$",bd=False)
put(n,"A9","IRR (annualised)",KEY)
put(n,"B9",f'=IFERROR((1+IRR(D13:{last}13))^12-1,"n/a")',BD,num="0.0%",f=GREEN)
put(n,"A10","Simple payback",KEY); put(n,"B10",'=IF(B4>0,B3/B4*12,"n/a")',BD,num="0.0",f=GREEN)
put(n,"C10","months",bd=False)
put(n,"A12","Month",BD,f=LIGHT); put(n,"A13","Net cash flow",BD,f=LIGHT); put(n,"A14","Cumulative",BD,f=LIGHT)
for i in range(NM_):
    cl = CL(4+i)
    put(n,f"{cl}12", i+1, num="0", al="center")
    put(n,f"{cl}13", f"=IF({cl}$12=1,-$B$3,0)+IF(AND({cl}$12>$B$6,{cl}$12<=$B$6+$B$7),$B$4/12,0)", num="#,##0")
    put(n,f"{cl}14", f"={cl}13" if i == 0 else f"={CL(3+i)}14+{cl}13", num="#,##0")
    n.column_dimensions[cl].width = 11
ch = LineChart(); ch.title = "Cumulative cash flow ($)"; ch.height = 8; ch.width = 24
ch.add_data(Reference(n, min_col=4, max_col=3+NM_, min_row=14, max_row=14), from_rows=True)
ch.set_categories(Reference(n, min_col=4, max_col=3+NM_, min_row=12, max_row=12))
n.add_chart(ch, "A17")

# ================================================================= SENSITIVITY
import json
sv = wb.create_sheet("Sensitivity"); sv.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [2,26,18,18,18,18,18]):
    sv.column_dimensions[col].width = w
band(sv, 1, 1, 7, "  SENSITIVITY   |   computed off the same mass-balance engine as MB Base / MB Sort", NAVY, H1)
put(sv,"B3","Values are static (generated with the model engine). Re-run build_wb.py to refresh "
            "after changing Dashboard inputs.", IT, bd=False)
r = 5
for title, hdr, rows in json.load(open("sens.json")):
    band(sv, r, 2, 2+len(hdr)-1, "  " + title); r += 1
    for j, h in enumerate(hdr): put(sv, f"{CL(2+j)}{r}", h, BD, f=LIGHT, al="center")
    r += 1
    for row in rows:
        for j, v in enumerate(row):
            nf = None
            if isinstance(v, float):
                h = hdr[j].lower()
                nf = ("0.0%" if ("pull" in h or "capture" in h or "dilution" in h)
                      else "#,##0" if ("$" in h or "oz" in h) else "0.000")
            c = put(sv, f"{CL(2+j)}{r}", v, num=nf, al="center" if j == 0 else None)
            if isinstance(v,(int,float)) and "npv" in hdr[j].lower():
                c.fill = fill(GREEN if v > 0 else "FCE4E4")
        r += 1
    r += 2
put(sv,f"B{r}","NPV breakeven on capacity capture is ~29%; cash breakeven ~21%.", KEY, bd=False)

wb.save(OUT)
print("built", OUT, "| dashboard rows:", X, "| out rows:", O, "| NETA row", NETA)
