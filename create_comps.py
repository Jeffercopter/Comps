import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLING
# ============================================================
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
data_font = Font(name='Calibri', size=10)
source_font = Font(name='Calibri', size=8, italic=True, color='808080')
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)
currency_format = '#,##0.0'
multiple_format = '0.0x'

def setup_sheet(ws, title, headers):
    ws.title = title
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def write_row(ws, row_idx, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

def format_number_cells(ws, row_idx, col_indices, fmt):
    for col in col_indices:
        cell = ws.cell(row=row_idx, column=col)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = fmt

# ============================================================
# SHEET 1: Equipment / Infrastructure Comparable Transactions
# ============================================================
ws1 = wb.active
headers1 = [
    'Target', 'Acquirer', 'Percentage\nacquired (%)', 'Country',
    'Completion\ndate', 'Revenue\n(A$m)', 'EBITDA\n(A$m)',
    'Enterprise Value\n(A$m)', 'EV/EBITDA\nMultiple', 'Source', 'Source Link'
]
setup_sheet(ws1, 'Equipment & Infrastructure', headers1)

equip_data = [
    # Row data: Target, Acquirer, %, Country, Date, Revenue, EBITDA, EV, EV/EBITDA, Source, Source Link
    [
        'SG Fleet Group Limited',
        'Pacific Equity Partners Pty Limited',
        '100%', 'Australia', 'Apr-25',
        1112.4, 171.0, 1400.0, 8.2,
        'ASX Scheme Booklet (Dec 2024); FY24 Operating EBITDA (historical); 100% EV (equity + net corporate debt); fleet financing debt excluded',
        'https://investors.sgfleet.com/DownloadFile.axd?file=/Report/ComNews/20241204/02890088.pdf'
    ],
    [
        'Benchmark Scaffolding Ltd.',
        'Acrow Limited',
        '100%', 'United Kingdom', 'Mar-24',
        9.0, 2.4, 10.0, 4.0,
        'ASX Announcement (1 Mar 2024); ~A$9M base + ~A$1M earnout; 3.8-4.0x stated range; historical EBITDA; upfront CFDF basis',
        'https://www.listcorp.com/asx/acf/acrow-limited/news/acquisition-of-benchmark-scaffolding-3001857.html'
    ],
    [
        'Pit N Portal Mining Services Pty Ltd',
        'Macmahon Holdings Limited',
        '100%', 'Australia', 'Feb-24',
        None, None, 10.0, None,
        'ASX Announcement (19 Dec 2023); A$10M asset sale (contracts & assets); standalone financials not disclosed',
        'https://www.marketindex.com.au/asx/mah/announcements/macmahon-completes-pit-n-portal-acquisition-6A1192683'
    ],
    [
        'MI Scaffold Pty Ltd',
        'Acrow Formwork and Construction Services Limited (nka: Acrow)',
        '100%', 'Australia', 'Nov-23',
        None, 6.6, 26.4, 4.0,
        'ASX Announcement (6 Nov 2023); A$26.4M upfront + A$9.9M deferred (A$36.3M total); ~4.0x upfront EV/EBITDA stated; EBITDA back-solved from stated multiple; historical basis',
        'https://www.listcorp.com/asx/acf/acrow-limited/news/acquisition-of-mi-scaffold-and-equity-raising-2952231.html'
    ],
    [
        'STA Traffic Management Pty Ltd',
        'AVADA Group Limited',
        '100%', 'Australia', 'Oct-23',
        None, None, 8.5, None,
        'ASX Announcement (24 Aug 2023); A$8.5M total consideration; EBITDA/Revenue not publicly disclosed',
        'https://www.marketscreener.com/quote/stock/AVADA-GROUP-LIMITED-129389843/news/AVADA-Group-Limited-entered-in-a-binding-agreement-to-acquire-Sta-Traffic-Management-Pty-Ltd-fir-AUD-44687162/'
    ],
    [
        'Business and assets of Construct Traffic Pty Ltd',
        'AVADA Group Limited',
        '100%', 'Australia', 'Aug-22',
        34.0, 5.0, 17.6, 3.5,
        'ASX Announcement (20 Jun 2022); A$17.6M upfront CFDF (up to A$23.5M with earnout); FY22 sustainable EBITDA ~A$5M (historical/normalised); asset acquisition',
        'https://www.listcorp.com/asx/avd/avada-group-limited/news/avada-confirms-completion-of-construct-traffic-acquisition-2754130.html'
    ],
    [
        'Karratha Machinery Hire',
        'SSH Group Limited',
        '100%', 'Australia', 'May-22',
        6.3, 3.8, 15.0, 3.9,
        'ASX Announcement (20 Apr 2022); A$15M total consideration; FY21 EBITDA ~$3.8M (historical)',
        'https://www.listcorp.com/asx/ssh/ssh-group-ltd/news/transformative-eps-accretive-equipment-business-acquisition-2698290.html'
    ],
    [
        'Pit N Portal Mining Services Pty Ltd / Pit N Portal Equipment',
        'Emeco Holdings Limited',
        '100%', 'Australia', 'Feb-20',
        101.0, 20.0, 72.0, 3.6,
        'ASX Announcement (29 Jan 2020); 3.6x FY19 Operating EBITDA explicitly stated (historical); A$62M cash + A$10M shares = A$72M total EV',
        'https://announcements.asx.com.au/asxpdf/20200129/pdf/44dl8skrl5by12.pdf'
    ],
    [
        'Uni-Span Australia Pty Limited',
        'Acrow Formwork and Construction Services Limited (nka: Acrow)',
        '100%', 'Australia', 'Oct-19',
        34.0, 4.8, 21.3, 4.4,
        'ASX Announcement (17 Oct 2019); A$21.25M upfront before earnouts; 4.4x FY19 normalised EBITDA stated (historical)',
        'https://www.listcorp.com/asx/acf/acrow-limited/news/acquisition-of-uni-span-australia-pty-ltd-2251229.html'
    ],
    [
        'Matilda Equipment Holdings Pty Ltd',
        'Emeco Holdings Limited',
        '100%', 'Australia', 'Jul-18',
        None, 24.2, 80.0, 3.3,
        'ASX Announcement (30 Apr 2018); A$80M total EV; ~3.3x annualised Operating EBITDA stated (historical)',
        'https://www.emecogroup.com/assets/reports/uploads/2018/07/1-20180430_Emeco-to-acquire-Matilda-Equipment.pdf'
    ],
]

for i, row in enumerate(equip_data, 2):
    write_row(ws1, i, row)
    # Format number columns: Revenue(6), EBITDA(7), EV(8)
    format_number_cells(ws1, i, [6, 7, 8], currency_format)
    # Format multiple column (9)
    cell = ws1.cell(row=i, column=9)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0"x"'

# Set column widths for Sheet 1
col_widths1 = [45, 45, 12, 15, 12, 12, 12, 15, 12, 55, 50]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 2: Staffing / Services Comparable Transactions
# ============================================================
ws2 = wb.create_sheet()
headers2 = [
    'Target', 'Acquirer', 'Percentage\nacquired (%)', 'Country',
    'Completion\ndate', 'Revenue\n(A$m)', 'EBITDA\n(A$m)',
    'Enterprise Value\n(A$m)', 'EV/EBITDA\nMultiple', 'Source', 'Source Link'
]
setup_sheet(ws2, 'Staffing & Services', headers2)

staff_data = [
    [
        'Infrawork Holdings Limited',
        'Peoplein Limited',
        '100%', 'New Zealand', 'Feb-26',
        None, 4.6, 22.1, 4.8,
        'ASX Announcement (20 Feb 2026); NZD 24M upfront CFDF (NZD 56M max); NZD 5M EBITDA (forward); 4.8x stated; converted at NZD/AUD ~0.92',
        'https://www.listcorp.com/asx/ppe/peoplein-limited/news/completion-of-acquisition-of-infrawork-holdings-3321455.html'
    ],
    [
        'First Choice Care / Edmen',
        'Healthcare Australia Pty Ltd.',
        '100%', 'Australia', 'Dec-25',
        None, 3.3, 20.3, 6.2,
        'ASX Announcement (Dec 2025); A$20.25M total; 6.2x FY25 earnings stated by PeopleIN (historical)',
        'https://www.businessnewsaustralia.com/articles/peoplein-to-sell-first-choice-care-and-edmen.html'
    ],
    [
        'Techforce Personnel Pty Limited',
        'ina.',
        '79%', 'Australia', 'Dec-25',
        None, None, 23.5, None,
        'ASX Announcement (27 Nov 2025); A$23.5M for 79.3% stake (~A$29.6M implied 100%); EBITDA/Revenue not disclosed',
        'https://www.listcorp.com/asx/ppe/peoplein-limited/news/peoplein-to-divest-techforce-personnel-3283656.html'
    ],
    [
        'APM Human Services International Limited',
        'Madison Dearborn Partners, LLC',
        '71%', 'Australia', 'Sep-24',
        2332.0, 281.0, 2100.0, 7.5,
        'ASX Scheme Booklet (Aug 2024); $1.45/share; ~A$1.3B equity + ~A$800M net debt = ~A$2.1B 100% EV; FY24 Adjusted EBITDA A$281M (historical); scheme of arrangement',
        'https://www.listcorp.com/asx/apm/apm-human-services-international/news/scheme-booklet-registered-with-asic-3070553.html'
    ],
    [
        'Owen Pacific Workforce Pty Ltd',
        'Ashley Services Group Limited',
        '100%', 'Australia', 'Feb-23',
        62.8, 4.5, 15.1, 3.4,
        'ASX Announcement (3 Feb 2023); A$14.2-15.9M range (midpoint used); EBITDA A$4.1-5.0M expected FY23 (forward); Revenue ~A$62.8M',
        'https://www.marketscreener.com/quote/stock/ASHLEY-SERVICES-GROUP-LIM-55288579/news/Ashley-Services-Completes-Acquisition-of-Owen-Pacific-Workforce-42906039/'
    ],
    [
        'Arbor E&T, LLC',
        'APM Human Services International Limited',
        '100%', 'United States', 'Nov-22',
        454.7, 50.0, 239.8, 4.8,
        'ASX Announcement (Nov 2022); A$239.8M (US$153.5M) total EV; FY22 EBITDA A$50M (historical)',
        'https://newsnreleases.com/2022/11/02/apm-completes-acquisition-of-equus-workforce-solutions/'
    ],
    [
        'Linc Personnel Pty Ltd',
        'Ashley Services Group Limited',
        '75%', 'Australia', 'Jul-22',
        14.0, 1.4, 5.6, 4.0,
        'ASX Announcement (30 May 2022); A$4.2M for 75% (implied 100% EV A$5.6M); normalised FY22 EBITDA A$1.4M (100% basis); historical',
        'https://www.listcorp.com/asx/ash/ashley-services-group/news/ashley-acquires-major-shareholding-in-linc-personnel-2730158.html'
    ],
    [
        'FIP Group Holdings Pty Ltd',
        'Peoplein Limited',
        '100%', 'Australia', 'Jun-22',
        60.4, 9.5, 45.0, 4.7,
        'ASX Announcement (3 Jun 2022); A$45M upfront CFDF (A$70M max with earnout); expected FY23 EBITDA ~A$9.5M (forward); 4.7x upfront',
        'https://www.businessnewsaustralia.com/articles/peoplein-buys-food-industry-people-for--45-million.html'
    ],
    [
        'Perigon Group Pty Ltd',
        'Peoplein Limited',
        '100%', 'Australia', 'Feb-22',
        None, 4.3, 16.0, 3.7,
        'ASX Announcement (9 Feb 2022); A$16M upfront CFDF (up to A$26.8M); expected FY23 EBITDA ~A$4.3M (forward); 3.7x upfront',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/Peoplein-Limited-agreed-to-acquire-Perigon-Group-Pty-Ltd-for-AUD-26-8-million-37827523/'
    ],
    [
        'Clustera Sverige AB',
        'APM Human Services International Limited',
        '100%', 'Sweden', 'Jan-22',
        10.6, 2.0, 15.0, 8.3,
        'ASX Announcement (29 Dec 2021); SEK 100M (~A$15M) total; expected FY22 EBITDA SEK 12M (~A$2M) (forward); ~8.3x implied',
        'https://www.listcorp.com/asx/apm/apm-human-services-international/news/apm-completes-ipo-acquisitions-and-amp-swedish-market-entry-2652618.html'
    ],
    [
        'GMT Group Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Dec-21',
        None, None, None, None,
        'ASX Announcement (Dec 2021); financial details not disclosed; described as not material to FY21 earnings',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-acquired-GMT-People-37179876/'
    ],
    [
        'Inverse Group Pty. Ltd.',
        'Hiremii Limited',
        '100%', 'Australia', 'Nov-21',
        12.0, None, 1.5, None,
        'ASX Announcement (Nov 2021); up to A$1.5M; revenue milestones A$12-16M; EBITDA not disclosed',
        'https://www.listcorp.com/asx/hmi/hiremii-limited/news/acquisition-of-inverse-group-2625801.html'
    ],
    [
        'Vision Surveys (QLD) Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '75%', 'Australia', 'Jul-21',
        None, None, 6.7, None,
        'ASX Announcement (1 Jun 2021); A$6.7M upfront for 75% (up to A$10.5M for 100%); combined EBITDA with Techforce ~A$5.8M FY22; 14% revenue CAGR',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-completed-the-acquisition-of-75-majority-stake-in-Vision-Surveys-Pty-Ltd-36025964/'
    ],
    [
        'Techforce Personnel Pty Limited',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '79%', 'Australia', 'Jun-21',
        None, 5.8, 30.0, 5.2,
        'ASX Announcement (1 Jun 2021); A$23.8M for 79.3% debt-free (implied 100% EV A$30.0M); expected FY22 EBITDA ~A$5.8M (100% basis); forward',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-entered-into-binding-agreements-to-acquire-Techforce-Personnel-Pty-Ltd-for-35497394/'
    ],
    [
        'Business of SwingShift Nurses',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Mar-21',
        None, 1.0, 3.1, 3.1,
        'ASX Announcement (15 Mar 2021); A$3.1M cash upfront; expected EBITDA A$1M over next 12 months (forward); ~3.1x implied',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-entered-into-a-binding-agreement-to-acquire-Business-of-SwingShift-Nurses-33481796/'
    ],
    [
        'Oncontractor Pty. Ltd.',
        'Hiremii Limited',
        '100%', 'Australia', 'Feb-21',
        None, None, None, None,
        'ASX Announcement (19 Feb 2021); acquired via share swap (14.3M shares); EBITDA/Revenue not disclosed',
        'https://www.listcorp.com/asx/hmi/hiremii-limited/news/hiremii-completes-acquisition-of-inverse-group-2635188.html'
    ],
    [
        'Assets of ECT4Health Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Feb-21',
        None, None, None, None,
        'ASX Announcement (Feb 2021); expected EBITDA A$1M; purchase price not separately disclosed',
        'https://www.listcorp.com/asx/ppe/peoplein-limited/news/completion-of-swingshift-nurses-2525245.html'
    ],
    [
        'Ecareer Employment Services Pty Ltd / Illuminate Search and Consulting',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Jan-21',
        None, 1.3, 5.2, 4.0,
        'ASX Announcement (17 Dec 2020); A$5.15M (~A$5.2M net of cash); expected EBITDA A$1.3M (forward); ~4.0x implied',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-completed-the-acquisition-of-Ecareer-Employment-Services-Pty-Ltd-Illuminat-33544375/'
    ],
    [
        'The Instruction Company Pty Ltd',
        'Ashley Services Group Limited',
        '100%', 'Australia', 'Sep-20',
        None, None, None, None,
        'ASX Announcement (15 Sep 2020); national rail training provider; financial terms not publicly disclosed',
        'https://ashleyservicesgroup.com.au/ticrail-expands-into-western-australia/'
    ],
    [
        'Serendipity (WA) Pty Ltd (t/a APM)',
        'Madison Dearborn Partners, LLC',
        '55%', 'Australia', 'Jun-20',
        1100.0, 148.4, 1500.0, 10.1,
        'Media reports (Jun 2020); Serendipity (WA) = APM legal entity; MDP acquired 55% at ~A$1.5-1.6B implied 100% EV from Quadrant PE; FY20 EBITDA ~A$148M (historical)',
        'https://michaelwest.com.au/apm-promoters-exit-jobless-profits-in-asx-float-turn-to-profiteering-from-disabled/'
    ],
    [
        'Cd Group Holdings Pty. Ltd.',
        'Ashley Services Group Limited',
        '80%', 'Australia', 'Dec-19',
        40.0, 4.1, 14.0, 3.4,
        'ASX Announcement (2019); A$11.2M for 80% (implied 100% EV A$14.0M); FY19 normalised EBITDA A$4.1M (100% basis); historical',
        'https://www.lexology.com/library/detail.aspx?g=fe4a314d-4853-4892-83f1-ce9fb0ba5d5b'
    ],
    [
        'First Choice Care Pty. Ltd.',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Jul-19',
        None, 3.4, 16.8, 4.9,
        'ASX Announcement (Jun 2019); A$16.8M combined upfront with Carestaff; expected FY20 EBITDA A$3.4M combined (forward)',
        'https://www.fool.com.au/2019/06/19/people-infrastructure-announces-more-acquisitions/'
    ],
    [
        'Carestaff Healthcare Solutions Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Jun-19',
        None, None, None, None,
        'ASX Announcement (Jun 2019); combined with First Choice Care (A$16.8M total); EBITDA A$3.4M combined; individual split not disclosed',
        'https://www.fool.com.au/2019/06/19/people-infrastructure-announces-more-acquisitions/'
    ],
    [
        'Halcyon Knights / Halcyon Knights Commercial and Contracting',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Jun-19',
        None, 3.9, 13.5, 3.5,
        'ASX Announcement (Jun 2019); A$13.5M upfront CFDF (up to A$21.75M with earnout); expected FY20 EBITDA A$3.9M (forward); ~3.5x upfront',
        'https://www.businessnewsaustralia.com/articles/people-infrastructure-buys-halcyon-knights-for--13-5m.html'
    ],
    [
        'Remaining Shares in Recon Solutions Pty Ltd and Project Partners Corp',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Mar-19',
        None, None, 2.8, None,
        'ASX Announcement (Mar 2019); A$2.8M for remaining 50% shares; combined EBITDA with VNS ~A$1.1M incremental',
        'https://www.businessnewsaustralia.com/articles/steve-scanlan-sells-remaining-stake-in-recon-group.html'
    ],
    [
        'Victorian Nurse Specialists Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Mar-19',
        None, None, 2.5, None,
        'ASX Announcement (Mar 2019); A$2.5M cash; combined EBITDA with Recon ~A$1.1M incremental; standalone EBITDA not disclosed',
        'https://www.finnewsnetwork.com.au/archives/finance_news_network222179.html'
    ],
    [
        'Network Nursing Agency Pty Ltd / Australian Healthcare Academy Pty Ltd',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Aug-18',
        None, 2.7, 8.0, 3.0,
        'ASX Announcement (Aug 2018); A$8M upfront (up to A$9.1M); FY19 EBITDA target ~A$2.65-2.8M (forward); ~3.0x upfront implied',
        'https://www.staffingindustry.com/news/global-daily-news/australia-people-infrastructure-acquires-nursing-agency-business'
    ],
    [
        '50% Stake in Recon Solutions Pty Ltd and 50% Stake in Recon Technology',
        'People Infrastructure Ltd (nka: Peoplein Limited)',
        '50%', 'Australia', 'Jan-18',
        None, 0.8, 2.8, 3.5,
        'ASX Announcement (Jan 2018); A$2.8M for 50% stake; expected A$0.8M EBITDA (50% share) over 12 months (forward); ~3.5x explicitly stated; consistent on 50% or 100% basis',
        'https://www.proactiveinvestors.com.au/companies/news/189851/people-infrastructure-s-acquisition-will-drive-revenue-higher-189851.html'
    ],
    [
        'AWX Pty Ltd',
        'People Infrastructure Pty Ltd (nka: Peoplein Limited)',
        '100%', 'Australia', 'Oct-16',
        None, None, 18.4, None,
        'MarketScreener; A$18.4M (A$16.6M cash + A$1.8M earnout); standalone EBITDA/Revenue not disclosed',
        'https://www.marketscreener.com/quote/stock/PEOPLEIN-LIMITED-38467017/news/People-Infrastructure-Ltd-acquired-AWX-Pty-Ltd-for-AUD-18-4-million-35119752/'
    ],
]

for i, row in enumerate(staff_data, 2):
    write_row(ws2, i, row)
    format_number_cells(ws2, i, [6, 7, 8], currency_format)
    cell = ws2.cell(row=i, column=9)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0"x"'

# Set column widths for Sheet 2
col_widths2 = [55, 50, 12, 15, 12, 12, 12, 15, 12, 55, 50]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SUMMARY ROWS
# ============================================================

# Sheet 1 summary
equip_multiples = [row[8] for row in equip_data if row[8] is not None]
if equip_multiples:
    summary_row = len(equip_data) + 3
    ws1.cell(row=summary_row, column=1, value='Summary Statistics').font = Font(name='Calibri', bold=True, size=10)

    ws1.cell(row=summary_row + 1, column=1, value='Mean EV/EBITDA').font = Font(name='Calibri', bold=True, size=10)
    mean_val = sum(equip_multiples) / len(equip_multiples)
    cell = ws1.cell(row=summary_row + 1, column=9, value=round(mean_val, 1))
    cell.number_format = '0.0"x"'
    cell.font = Font(name='Calibri', bold=True, size=10)

    ws1.cell(row=summary_row + 2, column=1, value='Median EV/EBITDA').font = Font(name='Calibri', bold=True, size=10)
    sorted_m = sorted(equip_multiples)
    n = len(sorted_m)
    if n % 2 == 0:
        median_val = (sorted_m[n//2 - 1] + sorted_m[n//2]) / 2
    else:
        median_val = sorted_m[n//2]
    cell = ws1.cell(row=summary_row + 2, column=9, value=round(median_val, 1))
    cell.number_format = '0.0"x"'
    cell.font = Font(name='Calibri', bold=True, size=10)

# Sheet 2 summary
staff_multiples = [row[8] for row in staff_data if row[8] is not None]
if staff_multiples:
    summary_row2 = len(staff_data) + 3
    ws2.cell(row=summary_row2, column=1, value='Summary Statistics').font = Font(name='Calibri', bold=True, size=10)

    ws2.cell(row=summary_row2 + 1, column=1, value='Mean EV/EBITDA').font = Font(name='Calibri', bold=True, size=10)
    mean_val2 = sum(staff_multiples) / len(staff_multiples)
    cell = ws2.cell(row=summary_row2 + 1, column=9, value=round(mean_val2, 1))
    cell.number_format = '0.0"x"'
    cell.font = Font(name='Calibri', bold=True, size=10)

    ws2.cell(row=summary_row2 + 2, column=1, value='Median EV/EBITDA').font = Font(name='Calibri', bold=True, size=10)
    sorted_m2 = sorted(staff_multiples)
    n2 = len(sorted_m2)
    if n2 % 2 == 0:
        median_val2 = (sorted_m2[n2//2 - 1] + sorted_m2[n2//2]) / 2
    else:
        median_val2 = sorted_m2[n2//2]
    cell = ws2.cell(row=summary_row2 + 2, column=9, value=round(median_val2, 1))
    cell.number_format = '0.0"x"'
    cell.font = Font(name='Calibri', bold=True, size=10)

# Freeze top row on both sheets
ws1.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'

# Save
wb.save('/home/user/Comps/Comparable_Transaction_Analysis.xlsx')
print("Workbook created successfully: Comparable_Transaction_Analysis.xlsx")
print(f"Sheet 1: {ws1.title} - {len(equip_data)} transactions")
print(f"Sheet 2: {ws2.title} - {len(staff_data)} transactions")
print(f"Equipment multiples available: {len(equip_multiples)} ({equip_multiples})")
print(f"Staffing multiples available: {len(staff_multiples)} ({staff_multiples})")
